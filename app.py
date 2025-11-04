import os
os.environ['OAUTHLIB_INSECURE_TRANSPORT'] = '1'

from dotenv import load_dotenv
from flask import Flask, render_template, request, session, redirect, url_for, jsonify, send_file
from flask_session import Session
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import Flow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from io import BytesIO
import tempfile, zipfile
import threading
import os
import shutil
import json

load_dotenv()

app = Flask(__name__, static_folder='static', static_url_path='/static')
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'dev-secret-key-change-in-production')

app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_PERMANENT'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)
app.config['SESSION_COOKIE_SECURE'] = False
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_NAME'] = 'gmail_oauth_session'
app.config['SESSION_REFRESH_EACH_REQUEST'] = True

Session(app)

SESSION_DIR = os.path.join(os.path.dirname(__file__), 'flask_session')
os.makedirs(SESSION_DIR, exist_ok=True)

# ✅ FIX: Job storage folder (persistent)
JOB_STORAGE_DIR = os.path.join(os.path.dirname(__file__), 'job_storage')
os.makedirs(JOB_STORAGE_DIR, exist_ok=True)

STATE_STORE = {}

SCOPES = [
    'https://www.googleapis.com/auth/gmail.send',
    'https://www.googleapis.com/auth/userinfo.email',
    'https://www.googleapis.com/auth/userinfo.profile',
    'https://www.googleapis.com/auth/gmail.readonly',
    'openid'
]

CLIENT_ID = os.environ.get('GOOGLE_CLIENT_ID')
CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET')
REDIRECT_URI = os.environ.get('REDIRECT_URI', 'http://localhost:5000/oauth2callback')

UPLOAD_FOLDERS = {}

if not CLIENT_ID or not CLIENT_SECRET:
    print("⚠️ WARNING: GOOGLE_CLIENT_ID or GOOGLE_CLIENT_SECRET not set!")
else:
    print("✅ OAuth Config Loaded Successfully")

# ✅ FIX: Job Status Helper Functions
def save_job_status(job_id, status_dict):
    """Lưu job status vào file"""
    job_file = os.path.join(JOB_STORAGE_DIR, f"{job_id}.json")
    with open(job_file, 'w', encoding='utf-8') as f:
        json.dump(status_dict, f)

def load_job_status(job_id):
    """Tải job status từ file"""
    job_file = os.path.join(JOB_STORAGE_DIR, f"{job_id}.json")
    if not os.path.exists(job_file):
        return None
    try:
        with open(job_file, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return None

def update_job_progress(job_id, current, total):
    """Cập nhật tiến độ job"""
    status = load_job_status(job_id)
    if status:
        status['progress'] = current
        status['total'] = total
        save_job_status(job_id, status)

def save_job_log(job_id, log_buffer):
    """Lưu log buffer vào file"""
    log_file = os.path.join(JOB_STORAGE_DIR, f"{job_id}_log.csv")
    log_buffer.seek(0)
    with open(log_file, 'wb') as f:
        f.write(log_buffer.read())

def load_job_log(job_id):
    """Tải log từ file"""
    log_file = os.path.join(JOB_STORAGE_DIR, f"{job_id}_log.csv")
    if os.path.exists(log_file):
        return log_file
    return None

@app.before_request
def make_session_permanent():
    session.permanent = True
    app.permanent_session_lifetime = timedelta(hours=24)

@app.route('/auth/login')
def oauth_login():
    try:
        print("\n🔵 [OAuth Login] Initiating...")
        
        flow = Flow.from_client_config(
            {
                "installed": {
                    "client_id": CLIENT_ID,
                    "client_secret": CLIENT_SECRET,
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [REDIRECT_URI]
                }
            },
            scopes=SCOPES,
            redirect_uri=REDIRECT_URI
        )
        
        auth_url, state = flow.authorization_url(
            access_type='offline',
            include_granted_scopes='true',
            prompt='consent'
        )
        
        session['oauth_state'] = state
        session['oauth_state_time'] = datetime.now().isoformat()
        STATE_STORE[state] = {
            'timestamp': datetime.now().isoformat(),
            'session_id': request.cookies.get('gmail_oauth_session', 'new')
        }
        session.modified = True
        
        print(f"✅ State created: {state[:40]}...")
        
        return redirect(auth_url)
        
    except Exception as e:
        print(f"❌ [OAuth Login] Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/oauth2callback')
def oauth_callback():
    try:
        print("\n🔵 [OAuth Callback] Received...")
        
        state_from_google = request.args.get('state')
        code = request.args.get('code')
        error = request.args.get('error')
        
        if error:
            print(f"❌ Google Error: {error}")
            return f"❌ Google Error: {error}", 400
        
        if not state_from_google or not code:
            print("❌ No state or code from Google")
            return "❌ Invalid OAuth response - Login again", 400
        
        state_from_session = session.get('oauth_state')
        state_from_memory = STATE_STORE.get(state_from_google, {})
        
        if not ((state_from_session and state_from_google == state_from_session) or state_from_memory):
            print("❌ State NOT verified - MISMATCH")
            return "❌ State verification failed - Login again", 400
        
        print(f"✅ State verification PASSED")
        
        flow = Flow.from_client_config(
            {
                "installed": {
                    "client_id": CLIENT_ID,
                    "client_secret": CLIENT_SECRET,
                    "auth_uri": "https://accounts.google.com/o/oauth2/auth",
                    "token_uri": "https://oauth2.googleapis.com/token",
                    "redirect_uris": [REDIRECT_URI]
                }
            },
            scopes=SCOPES,
            state=state_from_google,
            redirect_uri=REDIRECT_URI
        )
        
        flow.fetch_token(authorization_response=request.url)
        credentials = flow.credentials
        
        print(f"✅ Tokens received from Google")
        
        service = build('gmail', 'v1', credentials=credentials)
        profile = service.users().getProfile(userId='me').execute()
        user_email = profile.get('emailAddress', '')
        
        print(f"✅ User email: {user_email}")
        
        session['credentials'] = {
            'token': credentials.token,
            'refresh_token': credentials.refresh_token,
            'token_uri': credentials.token_uri,
            'client_id': credentials.client_id,
            'client_secret': credentials.client_secret,
            'scopes': credentials.scopes
        }
        session['user_email'] = user_email
        session['oauth_state'] = None
        session.modified = True
        
        if state_from_google in STATE_STORE:
            del STATE_STORE[state_from_google]
        
        print(f"✅ Session saved - Logged in as: {user_email}\n")
        
        return redirect(url_for('index'))
        
    except Exception as e:
        print(f"❌ [OAuth Callback] Exception: {str(e)}")
        import traceback
        traceback.print_exc()
        return f"❌ Error: {str(e)}", 500

@app.route('/auth/logout')
def oauth_logout():
    user_email = session.get('user_email', 'Unknown')
    print(f"\n✅ Logged out: {user_email}\n")
    session.clear()
    return redirect(url_for('index'))

@app.route('/')
def index():
    user_email = session.get('user_email')
    return render_template('index_multiuser.html', user_email=user_email)

@app.route('/get_sheets', methods=['POST'])
def get_sheets():
    try:
        file = request.files.get('file')
        if not file:
            return jsonify({'error': 'No file'}), 400
        
        file_bytes = file.read()
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=False)
        sheets = wb.sheetnames
        
        return jsonify({'sheets': sheets})
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/split', methods=['POST'])
def split_route():
    try:
        from modules.excel_splitter import split_excel_new
        return split_excel_new()
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return str(e), 500

@app.route('/upload_folder', methods=['POST'])
def upload_folder():
    try:
        files = request.files.getlist('files')
        if not files:
            return jsonify({'error': 'No files'}), 400
        
        session_id = session.sid or 'temp'
        temp_dir = tempfile.mkdtemp()
        UPLOAD_FOLDERS[session_id] = temp_dir
        
        for file in files:
            if file.filename.endswith('.xlsx'):
                file.save(os.path.join(temp_dir, file.filename))
        
        file_count = len(os.listdir(temp_dir))
        
        return jsonify({
            'success': True,
            'folder_id': session_id,
            'file_count': file_count
        })
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/send_emails', methods=['POST'])
def send_emails_route():
    try:
        if 'user_email' not in session:
            return jsonify({'error': 'Please login with Gmail first'}), 401
        
        if 'credentials' not in session:
            return jsonify({'error': 'OAuth token invalid - Please login again'}), 401
        
        print(f"\n🔵 [Send Emails] Started by: {session['user_email']}")
        
        job_id = datetime.now().strftime('%Y%m%d_%H%M%S')
        sender_name = request.form.get('sender_name', 'System')
        sender_email = session['user_email']
        
        folder_id = request.form.get('folder_id')
        email_file = request.files.get('email_file')
        
        ref_col = request.form['ref_col']
        name_col = request.form.get('name_col')
        email_col = request.form['email_col']
        cc_col = request.form.get('cc_col')
        subject = request.form['subject']
        body = request.form['body']
        
        start_row_email = int(request.form.get('start_row_email', 2))
        end_row_email = int(request.form.get('end_row_email', 99999))
        
        if not email_file or folder_id not in UPLOAD_FOLDERS:
            return jsonify({'error': 'Missing folder or email file'}), 400
        
        extract_folder = UPLOAD_FOLDERS[folder_id]
        
        temp_dir = tempfile.mkdtemp()
        excel_email_path = os.path.join(temp_dir, "emails.xlsx")
        email_file.save(excel_email_path)
        
        # ✅ FIX: Lưu job status vào file thay vì memory
        initial_status = {
            'status': 'processing',
            'progress': 0,
            'total': 0,
            'log_buffer': None
        }
        save_job_status(job_id, initial_status)
        
        creds_dict = session['credentials'].copy()
        
        def send_in_background():
            try:
                from modules.email_sender_oauth import send_emails_oauth
                
                credentials = Credentials(
                    token=creds_dict['token'],
                    refresh_token=creds_dict['refresh_token'],
                    token_uri=creds_dict['token_uri'],
                    client_id=creds_dict['client_id'],
                    client_secret=creds_dict['client_secret'],
                    scopes=creds_dict['scopes']
                )
                
                log_buffer = send_emails_oauth(
                    credentials=credentials,
                    sender_email=sender_email,
                    sender_name=sender_name,
                    excel_folder=extract_folder,
                    email_file_path=excel_email_path,
                    ref_col=ref_col,
                    name_col=name_col,
                    email_col=email_col,
                    cc_col=cc_col,
                    selected_col_for_match=ref_col,
                    subject_template=subject,
                    body_template=body,
                    start_row=start_row_email,
                    end_row=end_row_email,
                    progress_callback=lambda current, total: update_job_progress(job_id, current, total)
                )
                
                # ✅ FIX: Lưu log vào file
                save_job_log(job_id, log_buffer)
                
                # ✅ FIX: Cập nhật status thành completed
                status = load_job_status(job_id)
                status['status'] = 'completed'
                status['log_buffer'] = None  # Không lưu BytesIO trong JSON
                save_job_status(job_id, status)
                
                print(f"✅ [Send Emails] Completed")
            except Exception as e:
                status = load_job_status(job_id)
                status['status'] = 'failed'
                status['error'] = str(e)
                save_job_status(job_id, status)
                print(f"❌ [Send Emails] Failed: {str(e)}")
                import traceback
                traceback.print_exc()
        
        thread = threading.Thread(target=send_in_background, daemon=True)
        thread.start()
        
        return jsonify({
            'job_id': job_id,
            'message': 'Processing emails...'
        })
    
    except Exception as e:
        print(f"❌ [Send Emails] Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/check_status/<job_id>', methods=['GET'])
def check_status(job_id):
    # ✅ FIX: Load từ file thay vì memory
    status = load_job_status(job_id)
    if not status:
        return jsonify({'error': 'Job not found'}), 404
    
    return jsonify({
        'status': status['status'],
        'progress': status.get('progress', 0),
        'total': status.get('total', 0)
    })

@app.route('/download_log/<job_id>', methods=['GET'])
def download_log(job_id):
    # ✅ FIX: Load log từ file thay vì memory
    status = load_job_status(job_id)
    if not status:
        return "Job not found", 404
    
    if status['status'] != 'completed':
        return "Job not completed yet", 400
    
    log_file = load_job_log(job_id)
    if not log_file:
        return "No log available", 404
    
    filename = f"email_log_{job_id}.csv"
    return send_file(
        log_file, 
        as_attachment=True, 
        download_name=filename, 
        mimetype="text/csv"
    )

if __name__ == '__main__':
    app.run(debug=True, port=5000)