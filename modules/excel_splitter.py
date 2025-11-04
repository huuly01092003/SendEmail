from flask import request, send_file
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
import zipfile
import os

def column_letter_to_index(col_letter):
    """
    Convert column letter to index (1-based)
    Ví dụ: 'A' -> 1, 'B' -> 2, 'AA' -> 27
    """
    if not col_letter:
        return None
    col_letter = col_letter.upper().strip()
    result = 0
    for char in col_letter:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result

def split_excel_new():
    """
    Tách file Excel - FIX: Copy đúng header + data + chỉ các cột chọn
    """
    
    file = request.files.get('file')
    sheet_name = request.form.get('sheet_name')
    template_end_row = request.form.get('template_end_row')
    start_row = request.form.get('start_row')
    end_row = request.form.get('end_row')
    split_column = request.form.get('split_column')
    start_col = request.form.get('start_col', '').strip()
    end_col = request.form.get('end_col', '').strip()
    name_col = request.form.get('name_col')
    
    if not file or not sheet_name or not split_column or not template_end_row or not start_row or not end_row:
        return "❌ Missing required fields!", 400
    
    try:
        template_end_row = int(template_end_row)
        start_row = int(start_row)
        end_row = int(end_row)
    except ValueError:
        return "❌ Invalid row numbers!", 400
    
    try:
        file_bytes = file.read()
        
        # ✅ SỬA ĐỔI: Lấy tên file gốc và làm sạch nó
        original_filename_cleaned = "file_goc"
        if file and file.filename:
            # Tách tên, bỏ phần đuôi file (ví dụ: .xlsx)
            original_filename_cleaned = os.path.splitext(file.filename)[0]
            # Làm sạch tên file gốc (loại bỏ ký tự đặc biệt)
            original_filename_cleaned = original_filename_cleaned.replace("/", "_").replace("\\", "_").replace(":", "_")
            original_filename_cleaned = original_filename_cleaned.replace("*", "_").replace("?", "_").replace('"', "_")
            original_filename_cleaned = original_filename_cleaned.replace(" ", "") # Loại bỏ khoảng trắng cho gọn
        
        # Parse column range
        start_col_idx = None
        end_col_idx = None
        
        if start_col:
            start_col_idx = column_letter_to_index(start_col)
            if not start_col_idx:
                return f"❌ Invalid start column: {start_col}", 400
        
        if end_col:
            end_col_idx = column_letter_to_index(end_col)
            if not end_col_idx:
                return f"❌ Invalid end column: {end_col}", 400
        
        print(f"\n🔍 [Split] Sheet: {sheet_name}, Split Column: {split_column}")
        print(f"  Template End: {template_end_row}, Data: {start_row}-{end_row}")
        if start_col_idx or end_col_idx:
            print(f"  Column Range: {start_col or 'A'} to {end_col or 'Last'}")
        
        # Đọc file gốc (KHÔNG skip dòng nào)
        wb_orig = openpyxl.load_workbook(BytesIO(file_bytes))
        ws_orig = wb_orig[sheet_name]
        
        # Tìm header ở dòng template_end_row
        header_row_cells = []
        split_col_idx = None
        name_col_idx = None
        
        # Xác định max column cần đọc
        max_col_to_read = ws_orig.max_column
        if end_col_idx and end_col_idx < max_col_to_read:
            max_col_to_read = end_col_idx
        
        min_col_to_read = 1
        if start_col_idx and start_col_idx > 1:
            min_col_to_read = start_col_idx
        
        for col_idx in range(min_col_to_read, max_col_to_read + 1):
            cell_val = str(ws_orig.cell(template_end_row, col_idx).value).strip().lower()
            header_row_cells.append(cell_val)
            
            if cell_val == split_column.lower():
                split_col_idx = col_idx
            
            if name_col and cell_val == name_col.lower():
                name_col_idx = col_idx
        
        if split_col_idx is None:
            return f"❌ Column '{split_column}' not found! Available: {', '.join([c for c in header_row_cells if c != 'nan'])}", 400
        
        print(f"  ✅ Found split column at col {split_col_idx}")
        
        # Đọc data từ start_row đến end_row
        data_rows = []
        for row_idx in range(start_row, end_row + 1):
            # Kiểm tra xem dòng này có data không (check cột split)
            split_val = ws_orig.cell(row=row_idx, column=split_col_idx).value
            if split_val and str(split_val).strip() and str(split_val).lower() != 'nan':
                data_rows.append((split_val, row_idx))
        
        print(f"  ✅ Found {len(data_rows)} data rows")
        
        if len(data_rows) == 0:
            return "❌ No data found in range!", 400
        
        # Nhóm dữ liệu theo cột split
        grouped_data = {}
        for split_val, row_idx in data_rows:
            code = str(split_val).strip()
            if code not in grouped_data:
                grouped_data[code] = []
            grouped_data[code].append(row_idx)
        
        print(f"  ✅ Grouped into {len(grouped_data)} files")
        
        # Tạo ZIP
        zip_buffer = BytesIO()
        file_count = 0
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for code, row_indices in grouped_data.items():
                file_count += 1
                
                mã = code.replace("/", "_").replace("\\", "_").replace(":", "_")
                mã = mã.replace("*", "_").replace("?", "_").replace('"', "_")
                
                filename_part1 = mã # Tên file phần 1 (Mã)
                
                if name_col_idx and name_col_idx >= min_col_to_read:
                    row_idx_first = row_indices[0]
                    tên = str(ws_orig.cell(row=row_idx_first, column=name_col_idx).value).strip()
                    tên = tên.replace("/", "_").replace("\\", "_")
                    filename_part1 = f"{mã}-{tên}" # Tên file phần 1 (Mã-Tên)
                
                # ✅ SỬA ĐỔI: Nối tên file gốc vào
                final_filename = f"{filename_part1}-{original_filename_cleaned}"
                
                print(f"  [{file_count}] Creating: {final_filename}.xlsx")
                
                # Tạo file Excel
                buf = create_excel_file(
                    BytesIO(file_bytes),
                    ws_orig,
                    sheet_name,
                    template_end_row,
                    row_indices,
                    min_col_to_read,
                    max_col_to_read
                )
                
                # Sử dụng tên file cuối cùng
                zipf.writestr(f"{final_filename}.xlsx", buf.read())
        
        zip_buffer.seek(0)
        
        print(f"  ✅ Total files: {file_count}\n")
        
        return send_file(
            zip_buffer,
            as_attachment=True,
            download_name=f"tach_theo_{split_column}.zip",
            mimetype='application/zip'
        )
    
    except Exception as e:
        print(f"❌ Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return f"❌ Error: {str(e)}", 500

def create_excel_file(orig_file, ws_orig, sheet_name, template_end_row, row_indices, min_col, max_col):
    """
    Tạo file Excel: copy template + ghi data (FIX: chỉ copy cột cần thiết)
    
    :param row_indices: List các dòng data cần copy
    :param min_col: Cột bắt đầu (1-based index)
    :param max_col: Cột kết thúc (1-based index)
    """
    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    
    # Copy column widths (chỉ các cột được chọn)
    for col_idx in range(min_col, max_col + 1):
        col_letter = get_column_letter(col_idx)
        col_letter_new = get_column_letter(col_idx - min_col + 1)
        
        col_dim_orig = ws_orig.column_dimensions[col_letter]
        col_dim_new = ws_new.column_dimensions[col_letter_new]
        
        if col_dim_orig.width:
            col_dim_new.width = col_dim_orig.width
    
    # Copy row heights
    for row_num in range(1, template_end_row + 1):
        if row_num in ws_orig.row_dimensions:
            row_dim_orig = ws_orig.row_dimensions[row_num]
            row_dim_new = ws_new.row_dimensions[row_num]
            if row_dim_orig.height:
                row_dim_new.height = row_dim_orig.height
    
    # Copy template (dòng 1 đến template_end_row) - GIỮ NGUYÊN FORMAT
    for row_idx in range(1, template_end_row + 1):
        for col_idx in range(min_col, max_col + 1):
            cell_o = ws_orig.cell(row=row_idx, column=col_idx)
            cell_n = ws_new.cell(row=row_idx, column=col_idx - min_col + 1)
            
            cell_n.value = cell_o.value
            if cell_o.has_style:
                copy_cell_style(cell_o, cell_n)
    
    # Copy merged cells (chỉ những cells trong range cột)
    for merged_range in ws_orig.merged_cells.ranges:
        if merged_range.max_row <= template_end_row:
            # Check if merged range overlaps with our column range
            if merged_range.min_col <= max_col and merged_range.max_col >= min_col:
                try:
                    min_merged_col = max(merged_range.min_col, min_col)
                    max_merged_col = min(merged_range.max_col, max_col)
                    
                    min_merged_col_new = min_merged_col - min_col + 1
                    max_merged_col_new = max_merged_col - min_col + 1
                    
                    merged_range_new = f"{get_column_letter(min_merged_col_new)}{merged_range.min_row}:{get_column_letter(max_merged_col_new)}{merged_range.max_row}"
                    ws_new.merge_cells(merged_range_new)
                except:
                    pass
    
    # Ghi dữ liệu
    data_start_row = template_end_row + 1
    for idx, orig_row_idx in enumerate(row_indices):
        excel_row = data_start_row + idx
        
        # Copy row height
        if orig_row_idx in ws_orig.row_dimensions:
            row_dim = ws_orig.row_dimensions[orig_row_idx]
            if row_dim.height:
                ws_new.row_dimensions[excel_row].height = row_dim.height
        
        # Copy data từ row gốc (chỉ cột được chọn)
        for col_idx in range(min_col, max_col + 1):
            cell_o = ws_orig.cell(row=orig_row_idx, column=col_idx)
            cell_n = ws_new.cell(row=excel_row, column=col_idx - min_col + 1)
            
            cell_n.value = cell_o.value
            
            # Copy formatting từ ô dữ liệu gốc
            if cell_o.has_style:
                copy_cell_style(cell_o, cell_n)
    
    buffer = BytesIO()
    wb_new.save(buffer)
    buffer.seek(0)
    
    return buffer

def copy_cell_style(source, target):
    """Copy đầy đủ style"""
    if source.font:
        target.font = Font(
            name=source.font.name,
            size=source.font.size,
            bold=source.font.bold,
            italic=source.font.italic,
            color=source.font.color
        )
    
    if source.border:
        target.border = Border(
            left=source.border.left,
            right=source.border.right,
            top=source.border.top,
            bottom=source.border.bottom
        )
    
    if source.fill:
        target.fill = PatternFill(
            fill_type=source.fill.fill_type,
            start_color=source.fill.start_color,
            end_color=source.fill.end_color
        )
    
    if source.alignment:
        target.alignment = Alignment(
            horizontal=source.alignment.horizontal,
            vertical=source.alignment.vertical,
            wrap_text=source.alignment.wrap_text
        )
    
    if source.number_format:
        target.number_format = source.number_format