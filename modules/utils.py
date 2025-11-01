import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO

def copy_excel_template_and_insert_data_with_merged_cells(
    original_file_stream, group_df, template_end_row, data_start_row
):
    """
    Copy Excel template với support:
    - Merged cells
    - Formatting (màu sắc, borders, fonts)
    - Column width
    - Row height
    """
    original_file_stream.seek(0)
    workbook = openpyxl.load_workbook(original_file_stream)
    ws_original = workbook.active

    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active

    # ✅ 1️⃣ Copy cấu trúc cột (column widths) - FIX
    for col_letter in ws_original.column_dimensions:
        col_dim_orig = ws_original.column_dimensions[col_letter]
        col_dim_new = ws_new.column_dimensions[col_letter]
        
        # Copy width
        if col_dim_orig.width:
            col_dim_new.width = col_dim_orig.width
        
        # Copy hidden status
        if col_dim_orig.hidden:
            col_dim_new.hidden = col_dim_orig.hidden

    # ✅ 2️⃣ Copy cấu trúc dòng (row heights)
    for row_num in ws_original.row_dimensions:
        row_dim_orig = ws_original.row_dimensions[row_num]
        row_dim_new = ws_new.row_dimensions[row_num]
        
        if row_dim_orig.height:
            row_dim_new.height = row_dim_orig.height
        
        if row_dim_orig.hidden:
            row_dim_new.hidden = row_dim_orig.hidden

    # ✅ 3️⃣ Copy template (với formatting đầy đủ)
    for row_idx in range(1, template_end_row + 1):
        for col_idx in range(1, ws_original.max_column + 1):
            cell_o = ws_original.cell(row=row_idx, column=col_idx)
            cell_n = ws_new.cell(row=row_idx, column=col_idx)
            
            # Copy giá trị
            cell_n.value = cell_o.value
            
            # Copy formatting
            if cell_o.has_style:
                # Font
                if cell_o.font:
                    cell_n.font = openpyxl.styles.Font(
                        name=cell_o.font.name,
                        size=cell_o.font.size,
                        bold=cell_o.font.bold,
                        italic=cell_o.font.italic,
                        vertAlign=cell_o.font.vertAlign,
                        underline=cell_o.font.underline,
                        strike=cell_o.font.strike,
                        color=cell_o.font.color
                    )
                
                # Border
                if cell_o.border:
                    cell_n.border = openpyxl.styles.Border(
                        left=cell_o.border.left,
                        right=cell_o.border.right,
                        top=cell_o.border.top,
                        bottom=cell_o.border.bottom,
                        diagonal=cell_o.border.diagonal,
                        diagonal_direction=cell_o.border.diagonal_direction
                    )
                
                # Fill (màu nền)
                if cell_o.fill:
                    cell_n.fill = openpyxl.styles.PatternFill(
                        fill_type=cell_o.fill.fill_type,
                        start_color=cell_o.fill.start_color,
                        end_color=cell_o.fill.end_color
                    )
                
                # Alignment
                if cell_o.alignment:
                    cell_n.alignment = openpyxl.styles.Alignment(
                        horizontal=cell_o.alignment.horizontal,
                        vertical=cell_o.alignment.vertical,
                        text_rotation=cell_o.alignment.text_rotation,
                        wrap_text=cell_o.alignment.wrap_text,
                        shrink_to_fit=cell_o.alignment.shrink_to_fit,
                        indent=cell_o.alignment.indent
                    )
                
                # Number format
                if cell_o.number_format:
                    cell_n.number_format = cell_o.number_format
                
                # Protection
                if cell_o.protection:
                    cell_n.protection = openpyxl.styles.Protection(
                        locked=cell_o.protection.locked,
                        hidden=cell_o.protection.hidden
                    )

    # ✅ 4️⃣ Copy merged cells từ template
    for merged_cell_range in ws_original.merged_cells.ranges:
        # Chỉ copy merged cells trong template range
        min_row = merged_cell_range.min_row
        max_row = merged_cell_range.max_row
        
        if max_row <= template_end_row:
            ws_new.merge_cells(str(merged_cell_range))

    # ✅ 5️⃣ Ghi dữ liệu mới
    for r_idx, row in group_df.iterrows():
        excel_row_index = data_start_row + r_idx
        for c_idx, value in enumerate(row):
            cell = ws_new.cell(row=excel_row_index, column=c_idx + 1, value=value)
            
            # Copy formatting từ dòng template ngay phía trên
            if excel_row_index > template_end_row:
                template_row = template_end_row
                template_cell = ws_new.cell(row=template_row, column=c_idx + 1)
                
                if template_cell.has_style:
                    # Copy font
                    if template_cell.font:
                        cell.font = openpyxl.styles.Font(
                            name=template_cell.font.name or "Calibri",
                            size=template_cell.font.size or 11,
                            bold=template_cell.font.bold or False,
                            color=template_cell.font.color
                        )
                    
                    # Copy border
                    if template_cell.border:
                        cell.border = openpyxl.styles.Border(
                            left=template_cell.border.left,
                            right=template_cell.border.right,
                            top=template_cell.border.top,
                            bottom=template_cell.border.bottom
                        )
                    
                    # Copy fill
                    if template_cell.fill:
                        cell.fill = openpyxl.styles.PatternFill(
                            fill_type=template_cell.fill.fill_type,
                            start_color=template_cell.fill.start_color,
                            end_color=template_cell.fill.end_color
                        )
                    
                    # Copy alignment
                    if template_cell.alignment:
                        cell.alignment = openpyxl.styles.Alignment(
                            horizontal=template_cell.alignment.horizontal,
                            vertical=template_cell.alignment.vertical,
                            wrap_text=template_cell.alignment.wrap_text
                        )
                    
                    # Copy number format
                    if template_cell.number_format:
                        cell.number_format = template_cell.number_format

    # ✅ 6️⃣ Set row heights cho data rows
    for row_num in range(data_start_row, data_start_row + len(group_df)):
        if template_end_row in ws_new.row_dimensions:
            template_height = ws_new.row_dimensions[template_end_row].height
            if template_height:
                ws_new.row_dimensions[row_num].height = template_height

    # ✅ 7️⃣ Lưu workbook vào buffer
    buffer = BytesIO()
    wb_new.save(buffer)
    buffer.seek(0)
    
    return buffer


# ✅ HÀM CŨ - GIỮ LẠI CHO BACKWARD COMPATIBILITY
def copy_excel_template_and_insert_data(original_file_stream, group_df, template_end_row, data_start_row):
    """Hàm cũ - tương thích với code cũ"""
    return copy_excel_template_and_insert_data_with_merged_cells(
        original_file_stream, group_df, template_end_row, data_start_row
    )